"""
Serviço de exportação de dados para CSV e XLSX.
Suporta exportação de clientes, veículos, contratos e relatórios financeiros.
"""

from io import BytesIO
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import (
    Contrato,
    Cliente,
    Veiculo,
    DespesaVeiculo,
    DespesaLoja,
    Seguro,
    ParcelaSeguro,
    IpvaRegistro,
    Multa,
)


class ExportacaoService:
    """Serviço centralizado para exportação de dados em CSV e XLSX."""

    # Estilos XLSX
    HEADER_FILL = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _format_date(value: Optional[date]) -> str:
        """Formata data para DD/MM/AAAA."""
        if not value:
            return ""
        if isinstance(value, str):
            return value
        return value.strftime("%d/%m/%Y") if value else ""

    @staticmethod
    def _format_currency(value: Optional[float]) -> str:
        """Formata valor monetário com R$ e 2 casas decimais."""
        if value is None or value == "":
            return "R$ 0,00"
        try:
            num = float(value)
            return f"R$ {num:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
        except (ValueError, TypeError):
            return "R$ 0,00"

    @staticmethod
    def _format_percentage(value: Optional[float]) -> str:
        """Formata valor percentual com 2 casas decimais."""
        if value is None or value == "":
            return "0,00%"
        try:
            num = float(value)
            return f"{num:.2f}%".replace(".", ",")
        except (ValueError, TypeError):
            return "0,00%"

    @staticmethod
    def _create_csv_content(headers: List[str], rows: List[List], separator: str = ";") -> bytes:
        """Cria conteúdo CSV com BOM UTF-8."""
        content = "\ufeff"  # BOM UTF-8
        content += separator.join(f'"{h}"' for h in headers) + "\n"

        for row in rows:
            line = []
            for cell in row:
                cell_str = str(cell) if cell is not None else ""
                # Escapa aspas dentro do valor
                cell_str = cell_str.replace('"', '""')
                line.append(f'"{cell_str}"')
            content += separator.join(line) + "\n"

        return content.encode("utf-8")

    @staticmethod
    def _create_xlsx_workbook(
        sheet_name: str,
        headers: List[str],
        rows: List[List],
    ) -> openpyxl.Workbook:
        """Cria workbook XLSX com formatação padrão."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = ExportacaoService.HEADER_FILL
            cell.font = ExportacaoService.HEADER_FONT
            cell.alignment = ExportacaoService.HEADER_ALIGNMENT

        # Dados
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-width
        for col_num, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(
                len(str(header)) + 2, 12
            )

        return wb

    @staticmethod
    def _create_multi_sheet_xlsx(sheets_data: List[Tuple[str, List[str], List[List]]]) -> bytes:
        """Cria workbook XLSX com múltiplas abas."""
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        for sheet_name, headers, rows in sheets_data:
            ws = wb.create_sheet(sheet_name)

            # Cabeçalhos
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = ExportacaoService.HEADER_FILL
                cell.font = ExportacaoService.HEADER_FONT
                cell.alignment = ExportacaoService.HEADER_ALIGNMENT

            # Dados
            for row_num, row in enumerate(rows, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-width
            for col_num, header in enumerate(headers, 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)

        # Salva em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_clientes(
        db: Session,
        formato: str = "xlsx",
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
    ) -> BytesIO:
        """
        Exporta lista de clientes com informações de contratos.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            data_inicio: Filtro de data inicial de cadastro
            data_fim: Filtro de data final de cadastro

        Returns:
            BytesIO com dados do arquivo
        """
        query = db.query(Cliente)

        if data_inicio:
            query = query.filter(Cliente.data_cadastro >= data_inicio)
        if data_fim:
            query = query.filter(Cliente.data_cadastro <= data_fim)

        clientes = query.all()

        headers = [
            "ID",
            "Nome",
            "CPF/CNPJ",
            "RG",
            "Telefone",
            "E-mail",
            "CNH Número",
            "CNH Categoria",
            "CNH Validade",
            "Endereço",
            "Cidade",
            "Estado",
            "Score",
            "Data Cadastro",
            "Total Contratos",
            "Valor Total Locado",
        ]

        rows = []
        for cliente in clientes:
            # Contagem de contratos
            total_contratos = db.query(func.count(Contrato.id)).filter(
                Contrato.cliente_id == cliente.id
            ).scalar() or 0

            # Soma do valor total
            valor_total = db.query(func.sum(Contrato.valor_total)).filter(
                Contrato.cliente_id == cliente.id
            ).scalar() or 0

            row = [
                cliente.id,
                cliente.nome or "",
                cliente.cpf or "",
                cliente.rg or "",
                cliente.telefone or "",
                cliente.email or "",
                cliente.numero_cnh or "",
                cliente.categoria_cnh or "",
                ExportacaoService._format_date(cliente.validade_cnh),
                cliente.endereco_residencial or "",
                cliente.cidade_residencial or "",
                cliente.estado_residencial or "",
                cliente.score or "",
                ExportacaoService._format_date(cliente.data_cadastro),
                int(total_contratos),
                ExportacaoService._format_currency(valor_total),
            ]
            rows.append(row)

        if formato.lower() == "csv":
            content = ExportacaoService._create_csv_content(headers, rows)
            output = BytesIO(content)
        else:
            wb = ExportacaoService._create_xlsx_workbook("Clientes", headers, rows)
            output = BytesIO()
            wb.save(output)

        output.seek(0)
        return output

    @staticmethod
    def export_veiculos(
        db: Session,
        formato: str = "xlsx",
        status_filter: Optional[str] = None,
    ) -> BytesIO:
        """
        Exporta lista de veículos com informações de contratos e ROI.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            status_filter: Filtrar por status específico

        Returns:
            BytesIO com dados do arquivo
        """
        query = db.query(Veiculo)

        if status_filter:
            query = query.filter(Veiculo.status == status_filter)

        veiculos = query.all()

        headers = [
            "ID",
            "Placa",
            "Marca",
            "Modelo",
            "Ano",
            "Cor",
            "KM Atual",
            "Status",
            "Categoria",
            "Valor Diária",
            "Total Contratos",
            "Receita Total",
            "ROI (%)",
        ]

        rows = []
        for veiculo in veiculos:
            # Contagem de contratos
            total_contratos = db.query(func.count(Contrato.id)).filter(
                Contrato.veiculo_id == veiculo.id
            ).scalar() or 0

            # Receita total
            receita_total = db.query(func.sum(Contrato.valor_total)).filter(
                Contrato.veiculo_id == veiculo.id
            ).scalar() or 0

            # Cálculo ROI
            roi = 0
            valor_aquisicao = getattr(veiculo, 'valor_aquisicao', None) or 0
            if valor_aquisicao:
                try:
                    roi = (float(receita_total) / float(valor_aquisicao)) * 100
                except (ValueError, ZeroDivisionError):
                    roi = 0

            valor_diaria = veiculo.valor_diaria or 0
            if not valor_diaria and total_contratos > 0:
                avg_diaria = db.query(func.avg(Contrato.valor_diaria)).filter(
                    Contrato.veiculo_id == veiculo.id
                ).scalar()
                valor_diaria = avg_diaria or 0

            row = [
                veiculo.id,
                veiculo.placa or "",
                veiculo.marca or "",
                veiculo.modelo or "",
                veiculo.ano or "",
                veiculo.cor or "",
                veiculo.km_atual or 0,
                veiculo.status or "",
                getattr(veiculo, "categoria", "") or "",
                ExportacaoService._format_currency(valor_diaria),
                int(total_contratos),
                ExportacaoService._format_currency(receita_total),
                ExportacaoService._format_percentage(roi),
            ]
            rows.append(row)

        if formato.lower() == "csv":
            content = ExportacaoService._create_csv_content(headers, rows)
            output = BytesIO(content)
        else:
            wb = ExportacaoService._create_xlsx_workbook("Veículos", headers, rows)
            output = BytesIO()
            wb.save(output)

        output.seek(0)
        return output

    @staticmethod
    def export_contratos(
        db: Session,
        formato: str = "xlsx",
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
        status_filter: Optional[str] = None,
    ) -> BytesIO:
        """
        Exporta lista de contratos com detalhes de saída, devolução e custos.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            data_inicio: Filtro data inicial
            data_fim: Filtro data final
            status_filter: Filtrar por status

        Returns:
            BytesIO com dados do arquivo
        """
        query = db.query(Contrato).join(Cliente).join(Veiculo)

        if data_inicio:
            query = query.filter(Contrato.data_inicio >= data_inicio)
        if data_fim:
            query = query.filter(Contrato.data_inicio <= data_fim)
        if status_filter:
            query = query.filter(Contrato.status == status_filter)

        contratos = query.all()

        headers = [
            "ID",
            "Cliente",
            "CPF/CNPJ",
            "Veículo",
            "Data Saída",
            "Data Devolução Prevista",
            "Data Devolução Real",
            "KM Saída",
            "KM Retorno",
            "KM Percorrida",
            "Diárias",
            "Valor Diária",
            "Combustível Saída",
            "Combustível Retorno",
            "Avarias (R$)",
            "Desconto (R$)",
            "Valor Total (R$)",
            "Status",
            "Tipo",
        ]

        rows = []
        for contrato in contratos:
            km_percorrida = 0
            if contrato.km_inicial is not None and contrato.km_final is not None:
                km_percorrida = contrato.km_final - contrato.km_inicial

            # Cálculo de diárias
            diarias = contrato.qtd_diarias or 0
            if not diarias and contrato.data_inicio and contrato.data_fim:
                delta = contrato.data_fim - contrato.data_inicio
                diarias = delta.days or 1

            cliente_nome = contrato.cliente.nome if contrato.cliente else ""
            cliente_cpf = contrato.cliente.cpf if contrato.cliente else ""
            veiculo_info = ""
            if contrato.veiculo:
                veiculo_info = f"{contrato.veiculo.placa or ''} {contrato.veiculo.modelo or ''}".strip()

            row = [
                contrato.id,
                cliente_nome,
                cliente_cpf,
                veiculo_info,
                ExportacaoService._format_date(contrato.data_inicio),
                ExportacaoService._format_date(contrato.data_fim),
                ExportacaoService._format_date(contrato.data_fim),
                contrato.km_inicial or 0,
                contrato.km_final or 0,
                km_percorrida,
                diarias,
                ExportacaoService._format_currency(contrato.valor_diaria or 0),
                ExportacaoService._format_currency(contrato.combustivel_saida or 0),
                ExportacaoService._format_currency(contrato.combustivel_retorno or 0),
                ExportacaoService._format_currency(contrato.valor_avarias or 0),
                ExportacaoService._format_currency(contrato.desconto or 0),
                ExportacaoService._format_currency(contrato.valor_total or 0),
                contrato.status or "",
                contrato.tipo or "",
            ]
            rows.append(row)

        if formato.lower() == "csv":
            content = ExportacaoService._create_csv_content(headers, rows)
            output = BytesIO(content)
        else:
            wb = ExportacaoService._create_xlsx_workbook("Contratos", headers, rows)
            output = BytesIO()
            wb.save(output)

        output.seek(0)
        return output

    @staticmethod
    def export_financeiro(
        db: Session,
        formato: str = "xlsx",
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
    ) -> BytesIO:
        """
        Exporta relatório financeiro multi-aba com receitas, despesas e resumo.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            data_inicio: Filtro data inicial
            data_fim: Filtro data final

        Returns:
            BytesIO com dados do arquivo
        """
        sheets_data = []

        # ========== ABA 1: RESUMO MENSAL ==========
        resumo_headers = ["Mês", "Receita", "Despesa", "Lucro", "Margem %"]
        resumo_rows = ExportacaoService._generate_resumo_mensal(db, data_inicio, data_fim)
        sheets_data.append(("Resumo", resumo_headers, resumo_rows))

        # ========== ABA 2: RECEITAS ==========
        receitas_headers = [
            "ID",
            "Cliente",
            "CPF/CNPJ",
            "Veículo",
            "Data Saída",
            "Data Devolução Prevista",
            "Data Devolução Real",
            "KM Saída",
            "KM Retorno",
            "KM Percorrida",
            "Diárias",
            "Valor Diária",
            "Combustível Saída",
            "Combustível Retorno",
            "Avarias (R$)",
            "Desconto (R$)",
            "Valor Total (R$)",
            "Status",
            "Tipo",
        ]
        receitas_rows = ExportacaoService._generate_receitas(db, data_inicio, data_fim)
        sheets_data.append(("Receitas", receitas_headers, receitas_rows))

        # ========== ABA 3: DESPESAS VEÍCULOS ==========
        despesa_veiculo_headers = ["Veículo", "Tipo", "Descrição", "Data", "Valor"]
        despesa_veiculo_rows = ExportacaoService._generate_despesas_veiculos(db, data_inicio, data_fim)
        sheets_data.append(("Despesas Veículos", despesa_veiculo_headers, despesa_veiculo_rows))

        # ========== ABA 4: DESPESAS LOJA ==========
        despesa_loja_headers = ["Categoria", "Descrição", "Data", "Valor"]
        despesa_loja_rows = ExportacaoService._generate_despesas_loja(db, data_inicio, data_fim)
        sheets_data.append(("Despesas Loja", despesa_loja_headers, despesa_loja_rows))

        # ========== ABA 5: SEGUROS ==========
        seguros_headers = ["Apólice", "Vencimento", "Valor"]
        seguros_rows = ExportacaoService._generate_seguros(db, data_inicio, data_fim)
        sheets_data.append(("Seguros", seguros_headers, seguros_rows))

        # ========== ABA 6: IPVA ==========
        ipva_headers = ["Veículo", "Estado", "Valor", "Data Pagamento"]
        ipva_rows = ExportacaoService._generate_ipva(db, data_inicio, data_fim)
        sheets_data.append(("IPVA", ipva_headers, ipva_rows))

        # ========== ABA 7: MULTAS ==========
        multas_headers = ["Veículo", "Tipo", "Descrição", "Valor", "Data Pagamento"]
        multas_rows = ExportacaoService._generate_multas(db, data_inicio, data_fim)
        sheets_data.append(("Multas", multas_headers, multas_rows))

        if formato.lower() == "csv":
            # Para CSV, combina todas as abas com separadores
            content = "\ufeff"  # BOM UTF-8
            for sheet_name, headers, rows in sheets_data:
                content += f"\n{sheet_name}\n"
                content += ";".join(f'"{h}"' for h in headers) + "\n"
                for row in rows:
                    line = []
                    for cell in row:
                        cell_str = str(cell) if cell is not None else ""
                        cell_str = cell_str.replace('"', '""')
                        line.append(f'"{cell_str}"')
                    content += ";".join(line) + "\n"
                content += "\n"

            output = BytesIO(content.encode("utf-8"))
        else:
            # XLSX multi-aba
            xlsx_content = ExportacaoService._create_multi_sheet_xlsx(sheets_data)
            output = BytesIO(xlsx_content)

        output.seek(0)
        return output

    @staticmethod
    def _generate_resumo_mensal(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados do resumo mensal."""
        rows = []

        # Busca contratos finalizados
        query = db.query(Contrato).filter(Contrato.status == "finalizado")

        if data_inicio:
            query = query.filter(Contrato.data_fim >= data_inicio)
        if data_fim:
            query = query.filter(Contrato.data_fim <= data_fim)

        contratos = query.all()

        # Agrupa por mês
        meses = {}
        for contrato in contratos:
            if contrato.data_fim:
                chave_mes = contrato.data_fim.strftime("%m/%Y")
                if chave_mes not in meses:
                    meses[chave_mes] = {"receita": 0, "despesa": 0}
                meses[chave_mes]["receita"] += float(contrato.valor_total or 0)

        # Despesas por mês
        despesas_veiculo = db.query(DespesaVeiculo)
        if data_inicio:
            despesas_veiculo = despesas_veiculo.filter(DespesaVeiculo.data >= data_inicio)
        if data_fim:
            despesas_veiculo = despesas_veiculo.filter(DespesaVeiculo.data <= data_fim)

        for despesa in despesas_veiculo.all():
            if despesa.data:
                chave_mes = despesa.data.strftime("%m/%Y")
                if chave_mes not in meses:
                    meses[chave_mes] = {"receita": 0, "despesa": 0}
                meses[chave_mes]["despesa"] += float(despesa.valor or 0)

        despesas_loja = db.query(DespesaLoja)
        for despesa in despesas_loja.all():
            # DespesaLoja usa mes/ano como integers
            if hasattr(despesa, "mes") and hasattr(despesa, "ano"):
                chave_mes = f"{despesa.mes:02d}/{despesa.ano}"
                if chave_mes not in meses:
                    meses[chave_mes] = {"receita": 0, "despesa": 0}
                meses[chave_mes]["despesa"] += float(despesa.valor or 0)

        # Ordena por mês
        for chave_mes in sorted(meses.keys()):
            receita = meses[chave_mes]["receita"]
            despesa = meses[chave_mes]["despesa"]
            lucro = receita - despesa
            margem = (lucro / receita * 100) if receita > 0 else 0

            row = [
                chave_mes,
                ExportacaoService._format_currency(receita),
                ExportacaoService._format_currency(despesa),
                ExportacaoService._format_currency(lucro),
                ExportacaoService._format_percentage(margem),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_receitas(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de receitas (contratos finalizados)."""
        query = db.query(Contrato).filter(Contrato.status == "finalizado")

        if data_inicio:
            query = query.filter(Contrato.data_fim >= data_inicio)
        if data_fim:
            query = query.filter(Contrato.data_fim <= data_fim)

        contratos = query.all()

        rows = []
        for contrato in contratos:
            km_percorrida = 0
            if contrato.km_inicial is not None and contrato.km_final is not None:
                km_percorrida = contrato.km_final - contrato.km_inicial

            diarias = contrato.qtd_diarias or 0
            if not diarias and contrato.data_inicio and contrato.data_fim:
                delta = contrato.data_fim - contrato.data_inicio
                diarias = delta.days or 1

            cliente_nome = contrato.cliente.nome if contrato.cliente else ""
            cliente_cpf = contrato.cliente.cpf if contrato.cliente else ""
            veiculo_info = ""
            if contrato.veiculo:
                veiculo_info = f"{contrato.veiculo.placa or ''} {contrato.veiculo.modelo or ''}".strip()

            row = [
                contrato.id,
                cliente_nome,
                cliente_cpf,
                veiculo_info,
                ExportacaoService._format_date(contrato.data_inicio),
                ExportacaoService._format_date(contrato.data_fim),
                ExportacaoService._format_date(contrato.data_fim),
                contrato.km_inicial or 0,
                contrato.km_final or 0,
                km_percorrida,
                diarias,
                ExportacaoService._format_currency(contrato.valor_diaria or 0),
                ExportacaoService._format_currency(contrato.combustivel_saida or 0),
                ExportacaoService._format_currency(contrato.combustivel_retorno or 0),
                ExportacaoService._format_currency(contrato.valor_avarias or 0),
                ExportacaoService._format_currency(contrato.desconto or 0),
                ExportacaoService._format_currency(contrato.valor_total or 0),
                contrato.status or "",
                contrato.tipo or "",
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_despesas_veiculos(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de despesas com veículos."""
        query = db.query(DespesaVeiculo).join(Veiculo)

        if data_inicio:
            query = query.filter(DespesaVeiculo.data >= data_inicio)
        if data_fim:
            query = query.filter(DespesaVeiculo.data <= data_fim)

        despesas = query.all()

        rows = []
        for despesa in despesas:
            veiculo_info = ""
            if despesa.veiculo:
                veiculo_info = f"{despesa.veiculo.placa or ''} {despesa.veiculo.modelo or ''}".strip()

            row = [
                veiculo_info,
                despesa.tipo or "",
                despesa.descricao or "",
                ExportacaoService._format_date(despesa.data),
                ExportacaoService._format_currency(despesa.valor or 0),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_despesas_loja(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de despesas da loja."""
        despesas = db.query(DespesaLoja).all()

        rows = []
        for despesa in despesas:
            # Filtro manual por mes/ano se data_inicio/data_fim fornecidos
            if data_inicio or data_fim:
                if hasattr(despesa, "mes") and hasattr(despesa, "ano"):
                    despesa_date = date(despesa.ano, despesa.mes, 1)
                    if data_inicio and despesa_date < data_inicio:
                        continue
                    if data_fim and despesa_date > data_fim:
                        continue

            # Monta data a partir de mes/ano
            data_str = ""
            if hasattr(despesa, "mes") and hasattr(despesa, "ano"):
                data_str = f"{despesa.mes:02d}/{despesa.ano}"

            row = [
                despesa.categoria or "",
                despesa.descricao or "",
                data_str,
                ExportacaoService._format_currency(despesa.valor or 0),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_seguros(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de seguros (parcelas pagas)."""
        query = db.query(ParcelaSeguro).filter(ParcelaSeguro.status == "pago")

        if data_inicio:
            query = query.filter(ParcelaSeguro.vencimento >= data_inicio)
        if data_fim:
            query = query.filter(ParcelaSeguro.vencimento <= data_fim)

        parcelas = query.all()

        rows = []
        for parcela in parcelas:
            apólice = ""
            if parcela.seguro:
                apólice = parcela.seguro.numero_apolice or ""

            row = [
                apólice,
                ExportacaoService._format_date(parcela.vencimento),
                ExportacaoService._format_currency(parcela.valor or 0),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_ipva(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de IPVA."""
        query = db.query(IpvaRegistro)

        if data_inicio:
            query = query.filter(IpvaRegistro.data_pagamento >= data_inicio)
        if data_fim:
            query = query.filter(IpvaRegistro.data_pagamento <= data_fim)

        registros = query.all()

        rows = []
        for registro in registros:
            veiculo_info = ""
            if registro.veiculo:
                veiculo_info = f"{registro.veiculo.placa or ''} {registro.veiculo.modelo or ''}".strip()

            row = [
                veiculo_info,
                str(registro.ano_referencia or ""),
                ExportacaoService._format_currency(registro.valor_ipva or 0),
                ExportacaoService._format_date(registro.data_pagamento),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_multas(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de multas."""
        query = db.query(Multa)

        if data_inicio:
            query = query.filter(Multa.data_pagamento >= data_inicio)
        if data_fim:
            query = query.filter(Multa.data_pagamento <= data_fim)

        multas = query.all()

        rows = []
        for multa in multas:
            veiculo_info = ""
            if multa.veiculo:
                veiculo_info = f"{multa.veiculo.placa or ''} {multa.veiculo.modelo or ''}".strip()

            row = [
                veiculo_info,
                multa.tipo or "",
                multa.descricao or "",
                ExportacaoService._format_currency(multa.valor or 0),
                ExportacaoService._format_date(multa.data_pagamento),
            ]
            rows.append(row)

        return rows
