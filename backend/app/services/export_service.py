from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
import csv


class ExportService:
    """Service for exporting data to CSV and XLSX formats."""

    @staticmethod
    def export_contratos_xlsx(db: Session, filters: dict = None) -> BytesIO:
        """Export contracts to XLSX format."""
        from app.models import Contrato, Cliente, Veiculo

        wb = Workbook()
        ws = wb.active
        ws.title = "Contratos"

        # Headers
        headers = [
            "ID",
            "Número",
            "Cliente",
            "Veículo",
            "Data Início",
            "Data Fim",
            "Valor Diária",
            "Valor Total",
            "Status",
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Query contracts
        query = db.query(Contrato, Cliente, Veiculo).join(
            Cliente, Contrato.cliente_id == Cliente.id
        ).join(Veiculo, Contrato.veiculo_id == Veiculo.id)

        if filters:
            if filters.get("status"):
                query = query.filter(Contrato.status == filters["status"])

        contratos = query.all()

        # Add data
        for contrato, cliente, veiculo in contratos:
            ws.append(
                [
                    contrato.id,
                    contrato.numero,
                    cliente.nome,
                    f"{veiculo.marca} {veiculo.modelo} ({veiculo.placa})",
                    contrato.data_inicio.strftime("%d/%m/%Y %H:%M") if contrato.data_inicio else "",
                    contrato.data_fim.strftime("%d/%m/%Y %H:%M") if contrato.data_fim else "",
                    f"R$ {float(contrato.valor_diaria):.2f}",
                    f"R$ {float(contrato.valor_total):.2f}" if contrato.valor_total else "",
                    contrato.status,
                ]
            )

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 12

        # Format currency columns
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=8):
            for cell in row:
                cell.number_format = '"R$" #,##0.00'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_veiculos_xlsx(db: Session) -> BytesIO:
        """Export vehicles to XLSX format."""
        from app.models import Veiculo

        wb = Workbook()
        ws = wb.active
        ws.title = "Veículos"

        # Headers
        headers = [
            "ID",
            "Placa",
            "Marca",
            "Modelo",
            "Ano",
            "Cor",
            "Combustível",
            "KM Atual",
            "Status",
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Query vehicles
        veiculos = db.query(Veiculo).all()

        # Add data
        for veiculo in veiculos:
            ws.append(
                [
                    veiculo.id,
                    veiculo.placa,
                    veiculo.marca,
                    veiculo.modelo,
                    veiculo.ano,
                    veiculo.cor,
                    veiculo.combustivel,
                    f"{veiculo.km_atual:.2f}" if veiculo.km_atual else "",
                    veiculo.status,
                ]
            )

        # Adjust column widths
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col].width = 15

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_clientes_xlsx(db: Session) -> BytesIO:
        """Export clients to XLSX format."""
        from app.models import Cliente

        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Headers
        headers = [
            "ID",
            "Nome",
            "CPF",
            "Telefone",
            "Email",
            "Cidade",
            "Estado",
            "Score",
            "Ativo",
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Query clients
        clientes = db.query(Cliente).all()

        # Add data
        for cliente in clientes:
            ws.append(
                [
                    cliente.id,
                    cliente.nome,
                    cliente.cpf,
                    cliente.telefone or "",
                    cliente.email or "",
                    cliente.cidade_residencial or "",
                    cliente.estado_residencial or "",
                    cliente.score,
                    "Sim" if cliente.ativo else "Não",
                ]
            )

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_despesas_xlsx(db: Session, filters: dict = None) -> BytesIO:
        """Export expenses to XLSX format."""
        from app.models import DespesaContrato, DespesaVeiculo, DespesaLoja

        wb = Workbook()

        # Expense contracts sheet
        ws1 = wb.active
        ws1.title = "Despesas Contrato"

        headers1 = ["ID", "Contrato", "Tipo", "Descrição", "Valor", "Data"]
        ws1.append(headers1)

        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font

        despesas_contrato = db.query(DespesaContrato).all()
        for despesa in despesas_contrato:
            ws1.append(
                [
                    despesa.id,
                    despesa.contrato_id,
                    despesa.tipo or "",
                    despesa.descricao or "",
                    f"R$ {float(despesa.valor):.2f}" if despesa.valor else "",
                    despesa.data_registro.strftime("%d/%m/%Y") if despesa.data_registro else "",
                ]
            )

        # Expense vehicles sheet
        ws2 = wb.create_sheet("Despesas Veículo")
        headers2 = ["ID", "Veículo", "Descrição", "Valor", "KM", "Data"]
        ws2.append(headers2)

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        despesas_veiculo = db.query(DespesaVeiculo).all()
        for despesa in despesas_veiculo:
            ws2.append(
                [
                    despesa.id,
                    despesa.veiculo_id,
                    despesa.descricao or "",
                    f"R$ {float(despesa.valor):.2f}" if despesa.valor else "",
                    f"{despesa.km:.2f}" if despesa.km else "",
                    despesa.data.strftime("%d/%m/%Y") if despesa.data else "",
                ]
            )

        # Expense shop sheet
        ws3 = wb.create_sheet("Despesas Loja")
        headers3 = ["ID", "Mês", "Ano", "Descrição", "Valor", "Data"]
        ws3.append(headers3)

        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font

        despesas_loja = db.query(DespesaLoja).all()
        for despesa in despesas_loja:
            ws3.append(
                [
                    despesa.id,
                    despesa.mes,
                    despesa.ano,
                    despesa.descricao or "",
                    f"R$ {float(despesa.valor):.2f}" if despesa.valor else "",
                    despesa.data.strftime("%d/%m/%Y") if despesa.data else "",
                ]
            )

        # Adjust column widths for all sheets
        for ws in [ws1, ws2, ws3]:
            for col in ["A", "B", "C", "D", "E", "F"]:
                ws.column_dimensions[col].width = 15

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_contratos_csv(db: Session) -> BytesIO:
        """Export contracts to CSV format."""
        from app.models import Contrato, Cliente, Veiculo

        buffer = BytesIO()
        writer = None

        query = db.query(Contrato, Cliente, Veiculo).join(
            Cliente, Contrato.cliente_id == Cliente.id
        ).join(Veiculo, Contrato.veiculo_id == Veiculo.id)

        contratos = query.all()

        # Write CSV
        text_buffer = BytesIO()
        text_writer = None

        for contrato, cliente, veiculo in contratos:
            if text_writer is None:
                fieldnames = [
                    "ID",
                    "Número",
                    "Cliente",
                    "Veículo",
                    "Data Início",
                    "Data Fim",
                    "Valor Diária",
                    "Valor Total",
                    "Status",
                ]
                text_writer = csv.DictWriter(
                    text_buffer,
                    fieldnames=fieldnames,
                    lineterminator="\n",
                )
                text_writer.writeheader()

            text_writer.writerow(
                {
                    "ID": contrato.id,
                    "Número": contrato.numero,
                    "Cliente": cliente.nome,
                    "Veículo": f"{veiculo.marca} {veiculo.modelo}",
                    "Data Início": contrato.data_inicio.strftime("%d/%m/%Y %H:%M")
                    if contrato.data_inicio
                    else "",
                    "Data Fim": contrato.data_fim.strftime("%d/%m/%Y %H:%M")
                    if contrato.data_fim
                    else "",
                    "Valor Diária": f"R$ {float(contrato.valor_diaria):.2f}",
                    "Valor Total": f"R$ {float(contrato.valor_total):.2f}"
                    if contrato.valor_total
                    else "",
                    "Status": contrato.status,
                }
            )

        text_buffer.seek(0)
        return text_buffer
